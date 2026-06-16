import json
import boto3
import os
import csv
import io

from utils import log_operation, api_response

s3_client = boto3.client('s3')

BUCKET_NAME = os.environ['BUCKET_NAME']
LOGS_TABLE = os.environ['DYNAMODB_LOGS_TABLE']

ALLOWED_EXTENSIONS = ('.xlsx', '.xls')


def lambda_handler(event, context):
    """
    Converte file Excel (.xlsx / .xls) in CSV.

    NOTA: Richiede Lambda Layer con openpyxl installato.
    Vedi variabile Terraform: lambda_layer_arns_excel

    Input (body JSON):
    {
        "excel_key": "path/to/file.xlsx",
        "sheet_name": "Sheet1"  # opzionale, default primo foglio
    }
    """
    try:
        body = json.loads(event.get('body', '{}'))
        excel_key = body.get('excel_key')
        sheet_name = body.get('sheet_name', 0)  # Default: primo foglio

        if not excel_key:
            return api_response(400, {'error': 'excel_key is required'})

        # Valida estensione
        _, ext = os.path.splitext(excel_key.lower())
        if ext not in ALLOWED_EXTENSIONS:
            return api_response(400, {
                'error': f"Estensione non supportata: '{ext}'. Consentite: {ALLOWED_EXTENSIONS}"
            })

        # Scarica Excel da S3
        excel_obj = s3_client.get_object(Bucket=BUCKET_NAME, Key=excel_key)
        excel_data = excel_obj['Body'].read()

        try:
            import openpyxl
        except ImportError:
            error_msg = 'openpyxl non trovato. Aggiungere un Lambda Layer con openpyxl installato.'
            log_operation(LOGS_TABLE, 'excel_to_csv', {'error': error_msg}, 'error')
            return api_response(500, {
                'error': error_msg,
                'suggestion': 'Creare un layer con: pip install openpyxl -t python/ && zip -r layer.zip python/; poi impostare TF_VAR_lambda_layer_arns_excel=["arn:..."] e rieseguire terraform apply.'
            })

        workbook = openpyxl.load_workbook(io.BytesIO(excel_data), read_only=True, data_only=True)

        # Seleziona sheet
        if isinstance(sheet_name, str):
            if sheet_name not in workbook.sheetnames:
                return api_response(400, {
                    'error': f"Sheet '{sheet_name}' non trovato. Disponibili: {workbook.sheetnames}"
                })
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.worksheets[sheet_name]

        # Converti in CSV
        csv_buffer = io.StringIO()
        writer = csv.writer(csv_buffer)
        for row in sheet.iter_rows(values_only=True):
            writer.writerow(row)

        # Costruisci nome CSV sostituendo l'estensione originale
        csv_key = excel_key[:excel_key.rfind('.')] + '.csv'

        s3_client.put_object(
            Bucket=BUCKET_NAME,
            Key=csv_key,
            Body=csv_buffer.getvalue().encode('utf-8'),
            ContentType='text/csv'
        )

        log_operation(
            LOGS_TABLE,
            'excel_to_csv',
            {
                'excel_key': excel_key,
                'csv_key': csv_key,
                'sheet_name': str(sheet_name)
            }
        )

        return api_response(200, {
            'message': 'Excel convertito in CSV con successo',
            'csv_key': csv_key
        })

    except Exception as e:
        log_operation(LOGS_TABLE, 'excel_to_csv', {'error': str(e)}, 'error')
        return api_response(500, {'error': str(e)})
